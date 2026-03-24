import openpyxl

workbook = openpyxl.Workbook()

#get the default sheet
sheet=workbook["Sheet"]

#create a list of tuples as data source
data = [
    [225.7,'Gone with the Wind','Victor Fleming'],
    [194.4, 'Star Wars', 'George Lucas'],
    [161.0, 'ET: The Extraterrestrial', 'Steven Spielberg']
]
'''
for row in data:
    sheet.append(row)

#save the spreadsheet
workbook.save("movies.xlsx")
'''
'''
# i and j value to specify the row and column value for the excel file
for i, d in enumerate(data):
    for j, item in enumerate(d):
        sheet.cell(row=i+1, column=j+1).value = item
workbook.save("movies1.xlsx")
'''
i = 1
for eachrow in data:
    sheet.cell(row=i, column=1).value = eachrow[0]
    sheet.cell(row=i, column=2).value = eachrow[1]
    sheet.cell(row=i, column=3).value = eachrow[2]
    i += 1
workbook.save("data/movies2.xlsx")
    
    
    